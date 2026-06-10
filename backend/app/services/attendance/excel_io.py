from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from app.models.attendance import AttendanceRecordRead, AttendanceSessionRead


HEADER_ALIASES = {
    "student_code": {
        "student_code",
        "code",
        "code_etudiant",
        "cne",
        "cin",
        "apogee",
        "id",
        "identifiant",
        "identifiant_etudiant",
    },
    "first_name": {"first_name", "prenom", "prénom", "first name"},
    "last_name": {"last_name", "nom", "last name"},
    "group_name": {"group_name", "groupe", "group", "filiere", "filière"},
    "email": {"email", "mail", "adresse_email"},
}


class AttendanceImportRow:
    def __init__(
        self,
        student_code: str,
        first_name: str,
        last_name: str,
        group_name: str | None,
        email: str | None,
    ) -> None:
        self.student_code = student_code
        self.first_name = first_name
        self.last_name = last_name
        self.group_name = group_name
        self.email = email


class AttendanceExcelReader:
    def read_students(self, file_bytes: bytes) -> tuple[list[AttendanceImportRow], int]:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return [], 0

        headers = self._map_headers(rows[0])
        imported: list[AttendanceImportRow] = []
        skipped = 0

        for row in rows[1:]:
            item = self._row_to_item(row, headers)
            if item is None:
                skipped += 1
                continue
            imported.append(item)

        return imported, skipped

    def _map_headers(self, header_row: tuple[Any, ...]) -> dict[str, int]:
        mapped: dict[str, int] = {}

        for index, value in enumerate(header_row):
            normalized = self._normalize(value)
            for field_name, aliases in HEADER_ALIASES.items():
                if normalized in aliases and field_name not in mapped:
                    mapped[field_name] = index

        return mapped

    def _row_to_item(
        self,
        row: tuple[Any, ...],
        headers: dict[str, int],
    ) -> AttendanceImportRow | None:
        student_code = self._get_value(row, headers, "student_code")
        first_name = self._get_value(row, headers, "first_name")
        last_name = self._get_value(row, headers, "last_name")

        if not student_code or not first_name or not last_name:
            return None

        return AttendanceImportRow(
            student_code=student_code,
            first_name=first_name,
            last_name=last_name,
            group_name=self._get_value(row, headers, "group_name"),
            email=self._get_value(row, headers, "email"),
        )

    def _get_value(
        self,
        row: tuple[Any, ...],
        headers: dict[str, int],
        field_name: str,
    ) -> str | None:
        index = headers.get(field_name)
        if index is None or index >= len(row):
            return None

        value = row[index]
        if value is None:
            return None

        text = str(value).strip()
        return text or None

    def _normalize(self, value: Any) -> str:
        if value is None:
            return ""

        return str(value).strip().lower().replace(" ", "_").replace("-", "_")


class AttendanceExcelWriter:
    def write_session_export(
        self,
        session: AttendanceSessionRead,
        records: list[AttendanceRecordRead],
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"

        sheet.append(["Session", session.title])
        sheet.append(["Type", session.session_type])
        sheet.append(["Course", session.course_name or ""])
        sheet.append(["Group", session.group_name or ""])
        sheet.append([])
        sheet.append(
            [
                "Student Code",
                "First Name",
                "Last Name",
                "Group",
                "Status",
                "Recognized At",
                "Recognition Score",
            ]
        )

        for record in records:
            sheet.append(
                [
                    record.student_code,
                    record.first_name,
                    record.last_name,
                    record.group_name or "",
                    record.status,
                    record.recognized_at or "",
                    record.recognition_score if record.recognition_score is not None else "",
                ]
            )

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
